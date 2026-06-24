from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ReportWriter:
    """
    Writes analysis DataFrames to a single Excel workbook.

    Features
    --------
    - Creates workbook if it doesn't exist.
    - Replaces existing worksheet.
    - Formats header.
    - Auto fits column width.
    - Freezes header row.
    - Applies Auto Filter.
    """

    OUTPUT_FILE = Path("data/output/Reports/NBA_RCA_Report.xlsx")

    @staticmethod
    def write_sheet(dataframe: pd.DataFrame, sheet_name: str):
        """
        Write DataFrame to an Excel worksheet.
        """

        # Create output folder
        ReportWriter.OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

        # -------------------------
        # Create Workbook
        # -------------------------
        if not ReportWriter.OUTPUT_FILE.exists():

            with pd.ExcelWriter(
                ReportWriter.OUTPUT_FILE,
                engine="openpyxl"
            ) as writer:

                dataframe.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False
                )

        else:

            workbook = load_workbook(ReportWriter.OUTPUT_FILE)

            # Remove existing sheet
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

            workbook.save(ReportWriter.OUTPUT_FILE)
            workbook.close()

            with pd.ExcelWriter(
                ReportWriter.OUTPUT_FILE,
                engine="openpyxl",
                mode="a"
            ) as writer:

                dataframe.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False
                )

        # -------------------------
        # Apply Formatting
        # -------------------------

        workbook = load_workbook(ReportWriter.OUTPUT_FILE)
        worksheet = workbook[sheet_name]

        # Header Style
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # Format Header
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Freeze Header
        worksheet.freeze_panes = "A2"

        # Apply Filter
        worksheet.auto_filter.ref = worksheet.dimensions

        # Auto Fit Columns
        for column in worksheet.columns:

            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:

                try:
                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except Exception:
                    pass

            worksheet.column_dimensions[column_letter].width = min(
                max_length + 3,
                40
            )

        workbook.save(ReportWriter.OUTPUT_FILE)
        workbook.close()